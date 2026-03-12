from django import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.db.models import Case, Count, IntegerField, Q, Sum, When
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font
from django.views.decorators.http import require_GET, require_POST

from empleados_app.forms import AsignarDocenteForm, CarreraForm, CicloEscolarForm, CursoForm, EstablecimientoForm, GradoForm
from empleados_app.gafete_utils import resolve_gafete_dimensions
from empleados_app.models import Asistencia, AsistenciaDetalle, Carrera, CicloEscolar, ConfiguracionGeneral, Curso, CursoDocente, Empleado, Establecimiento, Grado, Matricula, Perfil, PeriodoAcademico
from empleados_app.permissions import (
    es_admin_total,
    es_docente,
    es_gestor,
    filtrar_por_establecimiento_usuario,
    obtener_establecimiento_usuario,
    usuario_puede_ver_establecimiento,
)

from .excel import autosize_columns, style_table_header, style_table_row, style_title, workbook_to_response
from .forms import MatriculaFiltroForm

ALLOW_MULTI_GRADE_PER_CYCLE = False


BASE_GAFETE_W = 1011
BASE_GAFETE_H = 639


def _canvas_for_orientation(orientation):
    return (BASE_GAFETE_W, BASE_GAFETE_H) if orientation == 'H' else (BASE_GAFETE_H, BASE_GAFETE_W)


def _resolve_gafete_dimensions(establecimiento, layout):
    orientation = str((layout or {}).get('canvas', {}).get('orientation') or ('V' if (establecimiento.gafete_alto or 0) > (establecimiento.gafete_ancho or 0) else 'H')).upper()
    if orientation not in ('H', 'V'):
        orientation = 'H'
    gafete_w, gafete_h = _canvas_for_orientation(orientation)
    return orientation, gafete_w, gafete_h


def _can_manage(user):
    return es_admin_total(user) or es_gestor(user)


def _is_docente(user):
    return es_docente(user)


def _can_view_attendance(user):
    return _is_docente(user) or _can_manage(user)


def _attendance_filter_for_user(user, prefix=""):
    if _is_docente(user):
        return {
            f"{prefix}docente": user,
            f"{prefix}curso__grado__carrera__ciclo_escolar__activo": True,
        }
    if es_gestor(user):
        establecimiento = obtener_establecimiento_usuario(user)
        if not establecimiento:
            return {f"{prefix}pk__in": []}
        return {f"{prefix}curso__grado__carrera__ciclo_escolar__establecimiento_id": establecimiento.id}
    return {}


def _ensure_establecimiento_access(request, est_id):
    if not usuario_puede_ver_establecimiento(request.user, est_id):
        messages.error(request, 'No tiene permisos para ese establecimiento.')
        return redirect('empleados:dahsboard')
    return None




def _gestores_qs_para_establecimiento(establecimiento):
    return (
        User.objects.filter(groups__name="Gestor", perfil__establecimiento_gestionado=establecimiento)
        .select_related("perfil")
        .order_by("first_name", "last_name", "username")
        .distinct()
    )


def _gestores_disponibles_qs():
    return User.objects.filter(groups__name="Gestor").order_by("first_name", "last_name", "username").distinct()


def _asignar_gestor_a_establecimiento(request, establecimiento):
    if not es_admin_total(request.user):
        messages.error(request, 'Solo un administrador puede gestionar asignaciones de gestores.')
        return

    action = (request.POST.get('action') or '').strip()
    gestor_id = request.POST.get('gestor_id')

    if not gestor_id:
        messages.warning(request, 'Debe seleccionar un gestor válido.')
        return

    gestor = get_object_or_404(_gestores_disponibles_qs(), pk=gestor_id)
    perfil, _ = Perfil.objects.get_or_create(user=gestor)

    if action == 'assign_gestor':
        anterior = perfil.establecimiento_gestionado
        perfil.establecimiento_gestionado = establecimiento
        perfil.save(update_fields=['establecimiento_gestionado'])
        if anterior and anterior.id != establecimiento.id:
            messages.success(request, f'Gestor reasignado desde "{anterior.nombre}" hacia "{establecimiento.nombre}".')
        else:
            messages.success(request, 'Gestor asignado correctamente al establecimiento.')
        return

    if action == 'unassign_gestor':
        if perfil.establecimiento_gestionado_id != establecimiento.id:
            messages.warning(request, 'El gestor seleccionado no está asignado a este establecimiento.')
            return
        perfil.establecimiento_gestionado = None
        perfil.save(update_fields=['establecimiento_gestionado'])
        messages.success(request, 'Gestor desasignado correctamente.')
        return

    messages.warning(request, 'Acción de gestor no reconocida.')

def _display_name_for_person(person):
    if not person:
        return ''

    get_full_name = getattr(person, 'get_full_name', None)
    if callable(get_full_name):
        full_name = (get_full_name() or '').strip()
        if full_name:
            return full_name

    nombres = getattr(person, 'nombres', '') or ''
    apellidos = getattr(person, 'apellidos', '') or ''
    nombre_empleado = f'{nombres} {apellidos}'.strip()
    if nombre_empleado:
        return nombre_empleado

    first_name = getattr(person, 'first_name', '') or ''
    last_name = getattr(person, 'last_name', '') or ''
    nombre_usuario = f'{first_name} {last_name}'.strip()
    if nombre_usuario:
        return nombre_usuario

    username = getattr(person, 'username', '') or ''
    if username:
        return username

    return str(person)




def _get_previous_cycle_for_establecimiento(ciclo_nuevo):
    return (
        CicloEscolar.objects.filter(establecimiento=ciclo_nuevo.establecimiento)
        .exclude(pk=ciclo_nuevo.pk)
        .order_by('-anio', '-id')
        .first()
    )


def _clone_academic_structure_from_previous_cycle(ciclo_nuevo):
    ciclo_anterior = _get_previous_cycle_for_establecimiento(ciclo_nuevo)
    if not ciclo_anterior:
        return {'copied': False, 'previous_cycle': None}

    carreras_anteriores = list(
        Carrera.objects.filter(ciclo_escolar=ciclo_anterior)
        .prefetch_related('grados__cursos')
        .order_by('id')
    )

    carrera_map = {}
    for carrera_anterior in carreras_anteriores:
        carrera_nueva, _ = Carrera.objects.get_or_create(
            ciclo_escolar=ciclo_nuevo,
            nombre=carrera_anterior.nombre,
            defaults={'activo': carrera_anterior.activo},
        )
        if carrera_nueva.activo != carrera_anterior.activo:
            carrera_nueva.activo = carrera_anterior.activo
            carrera_nueva.save(update_fields=['activo'])
        carrera_map[carrera_anterior.id] = carrera_nueva

    grado_map = {}
    for carrera_anterior in carreras_anteriores:
        carrera_nueva = carrera_map[carrera_anterior.id]
        for grado_anterior in carrera_anterior.grados.all().order_by('id'):
            grado_nuevo, _ = Grado.objects.get_or_create(
                carrera=carrera_nueva,
                nombre=grado_anterior.nombre,
                jornada=grado_anterior.jornada,
                seccion=grado_anterior.seccion,
                defaults={
                    'descripcion': grado_anterior.descripcion,
                    'activo': grado_anterior.activo,
                },
            )
            changed = False
            if grado_nuevo.descripcion != grado_anterior.descripcion:
                grado_nuevo.descripcion = grado_anterior.descripcion
                changed = True
            if grado_nuevo.activo != grado_anterior.activo:
                grado_nuevo.activo = grado_anterior.activo
                changed = True
            if changed:
                grado_nuevo.save(update_fields=['descripcion', 'activo'])
            grado_map[grado_anterior.id] = grado_nuevo

            for curso_anterior in grado_anterior.cursos.all().order_by('id'):
                curso_nuevo, _ = Curso.objects.get_or_create(
                    grado=grado_nuevo,
                    nombre=curso_anterior.nombre,
                    defaults={
                        'descripcion': curso_anterior.descripcion,
                        'activo': curso_anterior.activo,
                    },
                )
                curso_changed = False
                if curso_nuevo.descripcion != curso_anterior.descripcion:
                    curso_nuevo.descripcion = curso_anterior.descripcion
                    curso_changed = True
                if curso_nuevo.activo != curso_anterior.activo:
                    curso_nuevo.activo = curso_anterior.activo
                    curso_changed = True
                if curso_changed:
                    curso_nuevo.save(update_fields=['descripcion', 'activo'])

    return {'copied': True, 'previous_cycle': ciclo_anterior}

def _get_establecimiento(est_id):
    return get_object_or_404(Establecimiento, pk=est_id)


def _get_ciclo(est_id, ciclo_id):
    return get_object_or_404(CicloEscolar.objects.select_related('establecimiento'), pk=ciclo_id, establecimiento_id=est_id)

def _get_carrera(est_id, ciclo_id, car_id):
    return get_object_or_404(
        Carrera.objects.select_related('ciclo_escolar', 'ciclo_escolar__establecimiento'),
        pk=car_id,
        ciclo_escolar_id=ciclo_id,
        ciclo_escolar__establecimiento_id=est_id,
    )

def _get_carrera(est_id, ciclo_id, car_id):
    return get_object_or_404(
        Carrera.objects.select_related('ciclo_escolar', 'ciclo_escolar__establecimiento'),
        pk=car_id,
        ciclo_escolar_id=ciclo_id,
        ciclo_escolar__establecimiento_id=est_id,
    )




def _get_grado(est_id, ciclo_id, car_id, grado_id):
    return get_object_or_404(
        Grado.objects.select_related('carrera', 'carrera__ciclo_escolar', 'carrera__ciclo_escolar__establecimiento'),
        pk=grado_id,
        carrera_id=car_id,
        carrera__ciclo_escolar_id=ciclo_id,
        carrera__ciclo_escolar__establecimiento_id=est_id,
    )


@login_required
@user_passes_test(_can_manage)
def establecimientos_list(request):
    establecimientos = Establecimiento.objects.all()
    establecimientos = filtrar_por_establecimiento_usuario(establecimientos, request.user, 'id')
    return render(request, 'aulapro/establecimientos_list.html', {'establecimientos': establecimientos})


@login_required
@user_passes_test(_can_manage)
def establecimiento_detail(request, est_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied

    establecimiento = _get_establecimiento(est_id)

    if request.method == 'POST':
        _asignar_gestor_a_establecimiento(request, establecimiento)
        return redirect('empleados:establecimiento_detail', est_id=establecimiento.id)

    ciclos = establecimiento.ciclos_escolares.all().order_by('-anio', '-id')
    gestores_asignados = _gestores_qs_para_establecimiento(establecimiento)

    return render(request, 'aulapro/establecimiento_detail.html', {
        'establecimiento': establecimiento,
        'ciclos': ciclos,
        'ciclo_activo': establecimiento.get_ciclo_activo(),
        'gestores_asignados': gestores_asignados,
        'gestores_disponibles': _gestores_disponibles_qs(),
        'puede_gestionar_gestores': es_admin_total(request.user),
    })


@login_required
@user_passes_test(_can_manage)
def establecimiento_update(request, est_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    form = EstablecimientoForm(request.POST or None, request.FILES or None, instance=establecimiento)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Establecimiento actualizado correctamente.')
        return redirect('empleados:establecimiento_detail', est_id=establecimiento.id)

    return render(request, 'aulapro/establecimientos/form.html', {
        'establecimiento': establecimiento,
        'form': form,
        'titulo': 'Editar establecimiento',
    })


@login_required
@user_passes_test(_can_manage)
def ciclos_list(request, est_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclos = establecimiento.ciclos_escolares.all().order_by('-anio', '-id')
    return render(request, 'aulapro/ciclos_list.html', {
        'establecimiento': establecimiento,
        'ciclos': ciclos,
    })


@login_required
@user_passes_test(_can_manage)
def ciclo_create(request, est_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    form = CicloEscolarForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        ciclo = form.save(commit=False)
        ciclo.establecimiento = establecimiento
        ciclo.activo = bool(form.cleaned_data.get('activo'))

        try:
            with transaction.atomic():
                if ciclo.activo:
                    establecimiento.ciclos_escolares.filter(activo=True).exclude(pk=ciclo.pk).update(activo=False)
                ciclo.save()
                clone_result = _clone_academic_structure_from_previous_cycle(ciclo)
        except Exception:
            messages.error(request, 'No se pudo crear el ciclo escolar. Intente nuevamente.')
        else:
            if clone_result['copied']:
                messages.success(request, 'Ciclo creado correctamente. Se copió la estructura académica del ciclo anterior.')
            else:
                messages.success(request, 'Ciclo creado correctamente. No existía un ciclo anterior para copiar estructura.')
            return redirect('empleados:ciclo_detail', est_id=establecimiento.id, ciclo_id=ciclo.id)

    return render(request, 'aulapro/ciclos/form.html', {
        'establecimiento': establecimiento,
        'form': form,
        'titulo': 'Nuevo ciclo escolar',
        'ciclo': None,
        'accion': 'Guardar ciclo',
    })


@login_required
@user_passes_test(_can_manage)
def ciclo_update(request, est_id, ciclo_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = get_object_or_404(CicloEscolar, pk=ciclo_id, establecimiento=establecimiento)
    form = CicloEscolarForm(request.POST or None, instance=ciclo)

    if request.method == 'POST' and form.is_valid():
        ciclo = form.save(commit=False)
        ciclo.establecimiento = establecimiento
        ciclo.activo = bool(form.cleaned_data.get('activo'))

        with transaction.atomic():
            if ciclo.activo:
                establecimiento.ciclos_escolares.filter(activo=True).exclude(pk=ciclo.pk).update(activo=False)
            ciclo.save()

        messages.success(request, 'Ciclo escolar actualizado correctamente.')
        return redirect('empleados:ciclo_detail', est_id=establecimiento.id, ciclo_id=ciclo.id)

    return render(request, 'aulapro/ciclos/form.html', {
        'establecimiento': establecimiento,
        'form': form,
        'ciclo': ciclo,
        'titulo': 'Editar ciclo escolar',
        'accion': 'Guardar cambios',
    })


@login_required
@user_passes_test(_can_manage)
@require_POST
def ciclo_activar(request, est_id, ciclo_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = get_object_or_404(CicloEscolar, pk=ciclo_id, establecimiento=establecimiento)
    try:
        with transaction.atomic():
            establecimiento.ciclos_escolares.update(activo=False)
            ciclo.activo = True
            ciclo.save(update_fields=['activo'])
    except IntegrityError:
        messages.error(request, 'No se pudo activar el ciclo por un conflicto de integridad. Intente nuevamente.')
        return redirect('empleados:ciclo_detail', est_id=establecimiento.id, ciclo_id=ciclo.id)

    messages.success(request, f'El ciclo {ciclo.nombre} ahora está activo.')
    return redirect(request.POST.get('next') or 'empleados:ciclos_list', est_id=establecimiento.id)


@login_required
@user_passes_test(_can_manage)
@require_POST
def ciclo_delete(request, est_id, ciclo_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = get_object_or_404(CicloEscolar, pk=ciclo_id, establecimiento=establecimiento)

    if ciclo.matriculas.exists():
        messages.warning(request, 'No se puede eliminar el ciclo porque tiene matrículas asociadas.')
        return redirect('empleados:ciclo_detail', est_id=establecimiento.id, ciclo_id=ciclo.id)

    if ciclo.activo:
        messages.warning(request, 'No se puede eliminar el ciclo activo. Active otro ciclo primero.')
        return redirect('empleados:ciclo_detail', est_id=establecimiento.id, ciclo_id=ciclo.id)

    ciclo.delete()
    messages.success(request, 'Ciclo escolar eliminado correctamente.')
    return redirect('empleados:ciclos_list', est_id=establecimiento.id)



@login_required
@user_passes_test(_can_manage)
def ciclo_detail(request, est_id, ciclo_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carreras = ciclo.carreras.all().order_by('nombre')
    form = CarreraForm(initial={'ciclo_escolar': ciclo, 'activo': True})
    return render(request, 'aulapro/ciclo_detail.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carreras': carreras,
        'form_carrera': form,
    })




@login_required
@user_passes_test(_can_manage)
def carrera_create(request, est_id, ciclo_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    ciclo = _get_ciclo(est_id, ciclo_id)
    form = CarreraForm(request.POST or None, initial={'ciclo_escolar': ciclo, 'activo': True})
    if request.method == 'POST' and form.is_valid():
        carrera = form.save(commit=False)
        carrera.ciclo_escolar = ciclo
        carrera.save()
        messages.success(request, 'Carrera creada correctamente.')
        return redirect('empleados:ciclo_detail', est_id=est_id, ciclo_id=ciclo_id)

    return render(request, 'aulapro/carreras/form.html', {
        'establecimiento': ciclo.establecimiento,
        'ciclo': ciclo,
        'form': form,
        'titulo': 'Nueva carrera',
    })

@login_required
@user_passes_test(_can_manage)
def carrera_detail(request, est_id, ciclo_id, car_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grados = Grado.objects.filter(carrera=carrera).order_by('nombre')
    return render(request, 'aulapro/carrera_detail.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grados': grados,
    })


@login_required
@user_passes_test(_can_manage)
def grado_create(request, est_id, ciclo_id, car_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = get_object_or_404(Establecimiento, id=est_id)
    ciclo = get_object_or_404(CicloEscolar, id=ciclo_id, establecimiento=establecimiento)
    carrera = get_object_or_404(Carrera, id=car_id, ciclo_escolar=ciclo)
    grado = None

    form = GradoForm(request.POST or None, initial={'carrera': carrera, 'activo': True})
    if request.method == 'POST' and form.is_valid():
        grado = form.save(commit=False)
        grado.carrera = carrera
        grado.save()
        messages.success(request, 'Grado creado correctamente.')
        return redirect('empleados:carrera_detail', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id)

    return render(request, 'aulapro/grados/form.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'titulo': f'Nuevo grado - {carrera.nombre}',
        'form': form,
    })



@login_required
def carreras_list(request, est_id, ciclo_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    return ciclo_detail(request, est_id, ciclo_id)


@login_required
@user_passes_test(_can_manage)
def carrera_update(request, est_id, ciclo_id, car_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    form = CarreraForm(request.POST or None, instance=carrera)
    if request.method == 'POST' and form.is_valid():
        carrera = form.save(commit=False)
        carrera.ciclo_escolar = ciclo
        carrera.save()
        messages.success(request, 'Carrera actualizada correctamente.')
        return redirect('empleados:carrera_detail', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id)

    return render(request, 'aulapro/carreras/form.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'form': form,
        'titulo': 'Editar carrera',
    })


@login_required
@user_passes_test(_can_manage)
def grados_list(request, est_id, ciclo_id, car_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    return carrera_detail(request, est_id, ciclo_id, car_id)


@login_required
@user_passes_test(_can_manage)
def grado_update(request, est_id, ciclo_id, car_id, grado_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)
    form = GradoForm(request.POST or None, instance=grado)
    if request.method == 'POST' and form.is_valid():
        grado = form.save(commit=False)
        grado.carrera = carrera
        grado.save()
        messages.success(request, 'Grado actualizado correctamente.')
        return redirect('empleados:grado_detail', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id, grado_id=grado_id)

    return render(request, 'aulapro/grados/form.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'form': form,
        'titulo': 'Editar grado',
    })




@login_required
@user_passes_test(_can_manage)
@require_GET
def matricula_masiva_grado_buscar(request, est_id, ciclo_id, car_id, grado_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied

    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)
    establecimiento = grado.carrera.ciclo_escolar.establecimiento if grado.carrera and grado.carrera.ciclo_escolar else None
    if not establecimiento:
        return JsonResponse({'results': []}, status=400)

    q = (request.GET.get('q') or '').strip()
    if len(q) < 2:
        return JsonResponse({'results': []})

    alumnos = Empleado.objects.select_related('grado', 'establecimiento')

    # Reutiliza la lógica funcional de la matrícula masiva general:
    # una sola cadena `q` buscada en código, nombres o apellidos.
    alumnos = alumnos.filter(
        Q(codigo_personal__icontains=q)
        | Q(nombres__icontains=q)
        | Q(apellidos__icontains=q)
    )

    # En la matrícula por grado limitamos al establecimiento actual, sin excluir
    # alumnos que aún no tengan establecimiento asignado pero sí pertenezcan al mismo
    # establecimiento por su grado actual.
    alumnos = alumnos.filter(
        Q(establecimiento=establecimiento)
        | Q(establecimiento__isnull=True, grado__carrera__ciclo_escolar__establecimiento=establecimiento)
    ).order_by('codigo_personal', 'apellidos', 'nombres').distinct()[:15]

    results = [
        {
            'id': alumno.id,
            'codigo_personal': alumno.codigo_personal or '-',
            'nombre': f"{alumno.nombres} {alumno.apellidos}".strip(),
            'grado_actual': str(alumno.grado) if alumno.grado else '-',
            'establecimiento': str(alumno.establecimiento) if alumno.establecimiento else '-',
        }
        for alumno in alumnos
    ]
    return JsonResponse({'results': results})


@login_required
@user_passes_test(_can_manage)
def matricula_masiva_grado(request, est_id, ciclo_id, car_id, grado_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied

    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)

    ciclo_escolar = carrera.ciclo_escolar if carrera else None
    if not ciclo_escolar:
        messages.error(request, 'El grado no tiene ciclo escolar asociado. No se puede matricular masivamente.')
        return redirect('empleados:grado_detail', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id, grado_id=grado_id)

    establecimiento_usuario = obtener_establecimiento_usuario(request.user)
    if establecimiento_usuario and establecimiento_usuario.id != establecimiento.id:
        messages.error(request, 'No tiene permisos para matricular en este establecimiento.')
        return redirect('empleados:grado_detail', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id, grado_id=grado_id)

    if request.method == 'POST':
        raw_ids = request.POST.get('alumnos_ids', '')
        estado = request.POST.get('estado', 'activo')
        alumno_ids = [int(v) for v in raw_ids.split(',') if v.strip().isdigit()]
        alumno_ids = list(dict.fromkeys(alumno_ids))

        if not alumno_ids:
            messages.warning(request, 'Debe agregar al menos un alumno a la lista.')
            return render(request, 'aulapro/matricula_masiva_grado.html', {
                'establecimiento': establecimiento,
                'ciclo': ciclo,
                'carrera': carrera,
                'grado': grado,
                'ciclo_escolar_contexto': ciclo_escolar,
                'estados': Matricula.ESTADOS,
                'estado_default': estado,
            })

        alumnos_qs = Empleado.objects.filter(id__in=alumno_ids, establecimiento=establecimiento)
        alumnos_qs = filtrar_por_establecimiento_usuario(alumnos_qs, request.user, 'establecimiento_id')
        alumnos_map = {a.id: a for a in alumnos_qs}

        inscritos = 0
        omitidos = 0
        errores = 0

        for alumno_id in alumno_ids:
            alumno = alumnos_map.get(alumno_id)
            if not alumno:
                errores += 1
                continue

            if Matricula.objects.filter(alumno=alumno, grado=grado, ciclo_escolar=ciclo_escolar).exists():
                omitidos += 1
                continue

            matricula = Matricula(alumno=alumno, grado=grado, ciclo_escolar=ciclo_escolar, estado=estado)
            try:
                matricula.full_clean()
                matricula.save()
                inscritos += 1
            except ValidationError:
                errores += 1

        if inscritos:
            messages.success(request, f'{inscritos} alumnos matriculados correctamente.')
        if omitidos:
            messages.warning(request, f'{omitidos} alumnos ya tenían matrícula y fueron omitidos.')
        if errores:
            messages.error(request, f'{errores} alumnos no pudieron matricularse por validación o permisos.')

        return redirect('empleados:grado_detail', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id, grado_id=grado_id)

    return render(request, 'aulapro/matricula_masiva_grado.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'ciclo_escolar_contexto': ciclo_escolar,
        'estados': Matricula.ESTADOS,
        'estado_default': 'activo',
    })

@login_required
@user_passes_test(_can_manage)
def grado_detail(request, est_id, ciclo_id, car_id, grado_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)

    ciclo_activo = establecimiento.get_ciclo_activo()
    filtro_form = MatriculaFiltroForm(request.GET or None, establecimiento=establecimiento)
    matriculas = Matricula.objects.select_related('alumno', 'ciclo_escolar').filter(grado=grado)

    ciclo_filtrado = None
    if filtro_form.is_valid():
        estado = filtro_form.cleaned_data.get('estado')
        ciclo_filtrado = filtro_form.cleaned_data.get('ciclo_escolar')
        if estado:
            matriculas = matriculas.filter(estado=estado)

    if ciclo_filtrado:
        matriculas = matriculas.filter(ciclo_escolar=ciclo_filtrado)
    elif ciclo_activo:
        matriculas = matriculas.filter(ciclo_escolar=ciclo_activo)

    configuracion = ConfiguracionGeneral.objects.first()
    layout = establecimiento.get_layout()
    orientation, canvas_width, canvas_height = resolve_gafete_dimensions(establecimiento, layout)
    layout['canvas'] = {'width': canvas_width, 'height': canvas_height, 'orientation': orientation}
    return render(request, 'aulapro/grado_detail.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'matriculas': matriculas.order_by('-created_at', 'alumno__apellidos'),
        'filtro_form': filtro_form,
        'ciclo_activo': ciclo_activo,
        'configuracion': configuracion,
        'layout': layout,
        'canvas_width': canvas_width,
        'canvas_height': canvas_height,
        'gafete_w': canvas_width,
        'gafete_h': canvas_height,
        'orientacion': orientation,
    })


@login_required
@user_passes_test(_can_manage)
def cursos_list(request, est_id, ciclo_id, car_id, grado_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)
    cursos = Curso.objects.filter(grado=grado).order_by("nombre")
    return render(request, 'aulapro/grados/cursos/list.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'cursos': cursos,
    })


@login_required
@user_passes_test(_can_manage)
def curso_create(request, est_id, ciclo_id, car_id, grado_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)
    curso = None
    form = CursoForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        curso = form.save(commit=False)
        curso.grado = grado
        curso.save()
        messages.success(request, 'Curso creado correctamente.')
        return redirect('empleados:cursos_list', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id, grado_id=grado_id)
    return render(request, 'aulapro/grados/cursos/form.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'curso': curso,
        'titulo': 'Nuevo curso',
        'form': form,
    })


@login_required
@user_passes_test(_can_manage)
def curso_update(request, est_id, ciclo_id, car_id, grado_id, curso_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)
    curso = get_object_or_404(Curso, pk=curso_id, grado=grado)
    form = CursoForm(request.POST or None, instance=curso)
    if request.method == 'POST' and form.is_valid():
        curso = form.save(commit=False)
        curso.grado = grado
        curso.save()
        messages.success(request, 'Curso actualizado correctamente.')
        return redirect('empleados:cursos_list', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id, grado_id=grado_id)
    return render(request, 'aulapro/grados/cursos/form.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'curso': curso,
        'titulo': 'Editar curso',
        'form': form,
    })


@login_required
@user_passes_test(_can_manage)
def curso_asignar_docente(request, est_id, ciclo_id, car_id, grado_id, curso_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    establecimiento = _get_establecimiento(est_id)
    ciclo = _get_ciclo(est_id, ciclo_id)
    carrera = _get_carrera(est_id, ciclo_id, car_id)
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)
    curso = get_object_or_404(Curso, pk=curso_id, grado=grado)
    form = AsignarDocenteForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save(curso)
        messages.success(request, 'Docente asignado correctamente.')
        return redirect('empleados:cursos_list', est_id=est_id, ciclo_id=ciclo_id, car_id=car_id, grado_id=grado_id)
    asignaciones = CursoDocente.objects.filter(curso=curso).select_related('docente').order_by('docente__first_name', 'docente__last_name')
    return render(request, 'aulapro/grados/cursos/asignar_docente.html', {
        'establecimiento': establecimiento,
        'ciclo': ciclo,
        'carrera': carrera,
        'grado': grado,
        'curso': curso,
        'form': form,
        'asignaciones': asignaciones,
    })


@login_required
@user_passes_test(_can_view_attendance)
def docente_dashboard(request):
    cursos_docente = CursoDocente.objects.select_related(
        'curso', 'curso__grado', 'curso__grado__carrera', 'curso__grado__carrera__ciclo_escolar', 'curso__grado__carrera__ciclo_escolar__establecimiento', 'docente'
    ).filter(activo=True, curso__activo=True)

    show_docente_empty_message = False
    selected_ciclo = None
    ciclos_disponibles = CicloEscolar.objects.none()

    is_docente_user = _is_docente(request.user)

    if is_docente_user:
        cursos_docente = cursos_docente.filter(
            docente=request.user,
            curso__grado__carrera__ciclo_escolar__activo=True,
        )

        asignaciones_establecimientos = CursoDocente.objects.filter(
            docente=request.user,
            activo=True,
            curso__activo=True,
        ).values_list('curso__grado__carrera__ciclo_escolar__establecimiento_id', flat=True).distinct()

        if asignaciones_establecimientos and not CicloEscolar.objects.filter(
            establecimiento_id__in=asignaciones_establecimientos,
            activo=True,
        ).exists():
            messages.info(request, 'No existe un ciclo escolar activo.')
        elif not cursos_docente.exists():
            show_docente_empty_message = True
    else:
        ciclos_disponibles = CicloEscolar.objects.select_related('establecimiento').order_by('-activo', '-anio', '-id')
        ciclo_id = (request.GET.get('ciclo') or '').strip()
        if ciclo_id:
            selected_ciclo = ciclos_disponibles.filter(pk=ciclo_id).first()
        if selected_ciclo is None:
            selected_ciclo = ciclos_disponibles.filter(activo=True).first() or ciclos_disponibles.first()

        if selected_ciclo:
            cursos_docente = cursos_docente.filter(curso__grado__carrera__ciclo_escolar=selected_ciclo)
        else:
            cursos_docente = cursos_docente.none()
            messages.info(request, 'No existen ciclos escolares disponibles para filtrar.')

        if selected_ciclo and not cursos_docente.exists():
            messages.info(request, f'No hay cursos ni asistencias registradas para el ciclo "{selected_ciclo.nombre}".')

    cursos_docente = cursos_docente.order_by('curso__nombre')
    return render(request, 'docentes/dashboard.html', {
        'cursos_docente': cursos_docente,
        'show_docente_empty_message': show_docente_empty_message,
        'ciclos_disponibles': ciclos_disponibles,
        'selected_ciclo': selected_ciclo,
        'is_docente_user': is_docente_user,
    })


@login_required
@user_passes_test(_can_view_attendance)
def docente_curso_detail(request, curso_docente_id):
    curso_docente = get_object_or_404(
        CursoDocente.objects.select_related('curso', 'curso__grado', 'curso__grado__carrera', 'curso__grado__carrera__ciclo_escolar', 'curso__grado__carrera__ciclo_escolar__establecimiento'),
        pk=curso_docente_id,
        activo=True,
        **_attendance_filter_for_user(request.user),
    )
    grado = curso_docente.curso.grado
    alumnos = Empleado.objects.filter(matriculas__grado=grado).distinct().order_by('apellidos', 'nombres')
    return render(request, 'docentes/curso_detail.html', {
        'curso_docente': curso_docente,
        'curso': curso_docente.curso,
        'grado': grado,
        'alumnos': alumnos,
    })


@login_required
@user_passes_test(_can_view_attendance)
def docente_asistencia_home(request, curso_docente_id):
    curso_docente = get_object_or_404(CursoDocente.objects.select_related('curso', 'curso__grado'), pk=curso_docente_id, activo=True, **_attendance_filter_for_user(request.user))
    accion = request.GET.get('generar')
    if accion in {'bimestres', 'semestres'}:
        tipo, total = (PeriodoAcademico.TIPO_BIMESTRE, 4) if accion == 'bimestres' else (PeriodoAcademico.TIPO_SEMESTRE, 2)
        if PeriodoAcademico.objects.filter(curso_docente=curso_docente, tipo=tipo).exists():
            etiqueta = 'bimestre' if tipo == PeriodoAcademico.TIPO_BIMESTRE else 'semestre'
            messages.warning(request, f'Ya existe un {etiqueta} creado. Debe eliminarlo antes de crear uno nuevo.')
            return redirect('empleados:docente_asistencia_home', curso_docente_id=curso_docente.id)
        for i in range(1, total + 1):
            PeriodoAcademico.objects.create(
                curso_docente=curso_docente,
                tipo=tipo,
                numero=i,
                nombre=f"{tipo.title()} {i}",
                activo=True,
            )
        messages.success(request, 'Periodos generados correctamente.')
        return redirect('empleados:docente_asistencia_home', curso_docente_id=curso_docente.id)

    periodos = PeriodoAcademico.objects.filter(curso_docente=curso_docente).order_by('tipo', 'numero')
    return render(request, 'docentes/asistencia/home_periodos.html', {
        'curso_docente': curso_docente,
        'curso': curso_docente.curso,
        'periodos': periodos,
    })


@login_required
@user_passes_test(_can_view_attendance)
def docente_periodo_detail(request, periodo_id):
    periodo = get_object_or_404(PeriodoAcademico.objects.select_related('curso_docente', 'curso_docente__curso', 'curso_docente__curso__grado'), pk=periodo_id, **_attendance_filter_for_user(request.user, 'curso_docente__'))
    return render(request, 'docentes/asistencia/periodo_detail.html', {
        'periodo': periodo,
        'curso_docente': periodo.curso_docente,
        'curso': periodo.curso_docente.curso,
    })


@login_required
@user_passes_test(_can_view_attendance)
def tomar_asistencia(request, periodo_id):
    periodo = get_object_or_404(PeriodoAcademico.objects.select_related('curso_docente', 'curso_docente__curso', 'curso_docente__curso__grado'), pk=periodo_id, activo=True, **_attendance_filter_for_user(request.user, 'curso_docente__'))
    curso_docente = periodo.curso_docente
    fecha = request.POST.get('fecha') or request.GET.get('fecha') or str(timezone.localdate())
    grado = curso_docente.curso.grado
    alumnos = list(Empleado.objects.filter(matriculas__grado=grado).distinct().order_by('apellidos', 'nombres'))

    asistencia, _ = Asistencia.objects.get_or_create(curso_docente=curso_docente, periodo=periodo, fecha=fecha)

    for alumno in alumnos:
        AsistenciaDetalle.objects.get_or_create(asistencia=asistencia, alumno=alumno, defaults={'presente': True})

    detalles = list(AsistenciaDetalle.objects.filter(asistencia=asistencia).select_related('alumno').order_by('alumno__apellidos', 'alumno__nombres'))

    if request.method == 'POST':
        for detalle in detalles:
            detalle.presente = f'presente_{detalle.alumno_id}' in request.POST
        AsistenciaDetalle.objects.bulk_update(detalles, ['presente'])
        messages.success(request, 'Asistencia guardada correctamente.')
        return redirect('empleados:docente_historial_asistencias', periodo_id=periodo.id)

    return render(request, 'docentes/asistencia/tomar_asistencia.html', {
        'periodo': periodo,
        'curso_docente': curso_docente,
        'curso': curso_docente.curso,
        'fecha': fecha,
        'detalles': detalles,
    })


@login_required
@user_passes_test(_can_view_attendance)
def docente_historial_asistencias(request, periodo_id):
    periodo = get_object_or_404(PeriodoAcademico.objects.select_related('curso_docente', 'curso_docente__curso'), pk=periodo_id, **_attendance_filter_for_user(request.user, 'curso_docente__'))
    asistencias = Asistencia.objects.filter(periodo=periodo).order_by('-fecha')
    rows = []
    for a in asistencias:
        total = a.detalles.count()
        presentes = a.detalles.filter(presente=True).count()
        ausentes = total - presentes
        rows.append({'asistencia': a, 'total': total, 'presentes': presentes, 'ausentes': ausentes})
    return render(request, 'docentes/asistencia/historial.html', {
        'periodo': periodo,
        'curso_docente': periodo.curso_docente,
        'curso': periodo.curso_docente.curso,
        'rows': rows,
    })


@login_required
@user_passes_test(_can_view_attendance)
def docente_periodo_historial_excel(request, periodo_id):
    periodo = get_object_or_404(
        PeriodoAcademico.objects.select_related(
            'curso_docente',
            'curso_docente__docente',
            'curso_docente__curso',
            'curso_docente__curso__grado',
            'curso_docente__curso__grado__carrera',
            'curso_docente__curso__grado__carrera__ciclo_escolar',
            'curso_docente__curso__grado__carrera__ciclo_escolar__establecimiento',
        ),
        pk=periodo_id,
        **_attendance_filter_for_user(request.user, 'curso_docente__'),
    )

    curso_docente = periodo.curso_docente
    curso = curso_docente.curso
    grado = curso.grado

    alumnos = list(
        Empleado.objects.filter(matriculas__grado=grado)
        .distinct()
        .order_by('apellidos', 'nombres')
    )

    resumen_qs = AsistenciaDetalle.objects.filter(asistencia__periodo=periodo).values('alumno_id').annotate(
        asistencias=Sum(Case(When(presente=True, then=1), default=0, output_field=IntegerField())),
        inasistencias=Sum(Case(When(presente=False, then=1), default=0, output_field=IntegerField())),
        total_registros=Count('id'),
    )
    resumen_por_alumno = {
        row['alumno_id']: {
            'asistencias': row['asistencias'] or 0,
            'inasistencias': row['inasistencias'] or 0,
            'total_registros': row['total_registros'] or 0,
        }
        for row in resumen_qs
    }

    total_dias_registrados = Asistencia.objects.filter(periodo=periodo).count()
    establecimiento = '-'
    ciclo = '-'
    if curso.grado and curso.grado.carrera and curso.grado.carrera.ciclo_escolar:
        ciclo_obj = curso.grado.carrera.ciclo_escolar
        ciclo = ciclo_obj.nombre
        establecimiento = ciclo_obj.establecimiento.nombre

    wb = Workbook()
    ws = wb.active
    ws.title = 'Consolidado'

    style_title(ws, 1, 'Consolidado de asistencia del período', max_col=7)
    label_font = Font(bold=True)

    encabezado = [
        ('Curso:', curso.nombre),
        ('Docente:', _display_name_for_person(curso_docente.docente)),
        ('Período:', periodo.nombre),
        ('Ciclo escolar:', ciclo),
        ('Establecimiento:', establecimiento),
        ('Fecha de generación:', timezone.localdate().strftime('%d/%m/%Y')),
        ('Total de días con asistencia registrada:', total_dias_registrados),
    ]

    row = 3
    for label, value in encabezado:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    headers = ['No.', 'Código / carné', 'Alumno', 'Asistencias', 'Inasistencias', 'Total registros', '% Asistencia']
    style_table_header(ws, row, headers)
    ws.freeze_panes = f'A{row + 1}'

    for idx, alumno in enumerate(alumnos, start=1):
        resumen = resumen_por_alumno.get(alumno.id, {'asistencias': 0, 'inasistencias': 0, 'total_registros': 0})
        total_registros = resumen['total_registros']
        porcentaje = (resumen['asistencias'] / total_registros * 100) if total_registros else 0
        style_table_row(
            ws,
            row + idx,
            [
                idx,
                alumno.codigo_personal or '-',
                f'{alumno.apellidos}, {alumno.nombres}',
                resumen['asistencias'],
                resumen['inasistencias'],
                total_registros,
                f'{porcentaje:.2f}%',
            ],
        )

    autosize_columns(ws)
    return workbook_to_response(wb, f'consolidado_periodo_{periodo.id}')


@login_required
@user_passes_test(_can_view_attendance)
def docente_asistencia_detail(request, asistencia_id):
    asistencia = get_object_or_404(
        Asistencia.objects.select_related('curso_docente', 'curso_docente__docente', 'curso_docente__curso', 'periodo'),
        pk=asistencia_id,
        **_attendance_filter_for_user(request.user, 'curso_docente__'),
    )
    detalles = asistencia.detalles.select_related('alumno').order_by('alumno__apellidos', 'alumno__nombres')
    presentes = detalles.filter(presente=True).count()
    ausentes = detalles.count() - presentes
    return render(request, 'docentes/asistencia/detail.html', {
        'asistencia': asistencia,
        'periodo': asistencia.periodo,
        'curso_docente': asistencia.curso_docente,
        'curso': asistencia.curso_docente.curso,
        'detalles': detalles,
        'presentes': presentes,
        'ausentes': ausentes,
    })


@login_required
@user_passes_test(_can_view_attendance)
def docente_asistencia_excel(request, asistencia_id):
    asistencia = get_object_or_404(
        Asistencia.objects.select_related(
            'periodo', 'curso_docente', 'curso_docente__docente', 'curso_docente__curso',
            'curso_docente__curso__grado', 'curso_docente__curso__grado__carrera',
            'curso_docente__curso__grado__carrera__ciclo_escolar',
            'curso_docente__curso__grado__carrera__ciclo_escolar__establecimiento'
        ),
        pk=asistencia_id,
        **_attendance_filter_for_user(request.user, 'curso_docente__'),
    )
    detalles = list(asistencia.detalles.select_related('alumno').order_by('alumno__apellidos', 'alumno__nombres'))

    curso = asistencia.curso_docente.curso
    docente = asistencia.curso_docente.docente
    periodo = asistencia.periodo.nombre if asistencia.periodo else 'Sin periodo'
    ciclo = '-'
    establecimiento = '-'
    if curso.grado and curso.grado.carrera and curso.grado.carrera.ciclo_escolar:
        ciclo = curso.grado.carrera.ciclo_escolar.nombre
        establecimiento = curso.grado.carrera.ciclo_escolar.establecimiento.nombre

    wb = Workbook()
    ws = wb.active
    ws.title = 'Asistencia'

    style_title(ws, 1, 'Asistencia')

    encabezado = [
        ('Curso:', curso.nombre),
        ('Docente:', _display_name_for_person(docente)),
        ('Fecha:', asistencia.fecha.strftime('%d/%m/%Y')),
        ('Periodo:', periodo),
        ('Ciclo escolar:', ciclo),
        ('Establecimiento:', establecimiento),
    ]

    row = 3
    label_font = Font(bold=True)
    for label, value in encabezado:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    headers = ['No.', 'Código / Carné', 'Alumno', 'Estado', 'Observación']
    style_table_header(ws, row, headers)
    ws.freeze_panes = f'A{row + 1}'

    presentes = 0
    ausentes = 0
    for idx, detalle in enumerate(detalles, start=1):
        estado = 'Presente' if detalle.presente else 'Ausente'
        if detalle.presente:
            presentes += 1
        else:
            ausentes += 1
        style_table_row(
            ws,
            row + idx,
            [
                idx,
                detalle.alumno.codigo_personal or '-',
                f'{detalle.alumno.apellidos}, {detalle.alumno.nombres}',
                estado,
                '-',
            ],
        )

    total_row = row + len(detalles) + 2
    ws.cell(row=total_row, column=1, value='Total presentes').font = label_font
    ws.cell(row=total_row, column=2, value=presentes)
    ws.cell(row=total_row + 1, column=1, value='Total ausentes').font = label_font
    ws.cell(row=total_row + 1, column=2, value=ausentes)

    autosize_columns(ws)
    return workbook_to_response(wb, f'asistencia_{asistencia.fecha}')


@login_required
@user_passes_test(_can_view_attendance)
def docente_alumno_historial(request, curso_docente_id, alumno_id):
    curso_docente = get_object_or_404(CursoDocente.objects.select_related('curso', 'curso__grado'), pk=curso_docente_id, activo=True, **_attendance_filter_for_user(request.user))
    alumno = get_object_or_404(Empleado, pk=alumno_id)
    detalles = AsistenciaDetalle.objects.select_related('asistencia', 'asistencia__periodo').filter(
        asistencia__curso_docente=curso_docente,
        alumno=alumno,
    ).order_by('-asistencia__fecha')
    presentes = detalles.filter(presente=True).count()
    ausentes = detalles.count() - presentes

    resumen_periodos = {}
    for d in detalles:
        key = d.asistencia.periodo.nombre if d.asistencia.periodo else 'Sin periodo'
        resumen_periodos.setdefault(key, {'presentes': 0, 'ausentes': 0})
        if d.presente:
            resumen_periodos[key]['presentes'] += 1
        else:
            resumen_periodos[key]['ausentes'] += 1

    return render(request, 'docentes/alumno_historial.html', {
        'curso_docente': curso_docente,
        'curso': curso_docente.curso,
        'alumno': alumno,
        'detalles': detalles,
        'presentes': presentes,
        'ausentes': ausentes,
        'resumen_periodos': resumen_periodos,
    })


@login_required
@user_passes_test(_can_view_attendance)
def docente_alumno_historial_excel(request, curso_docente_id, alumno_id):
    curso_docente = get_object_or_404(
        CursoDocente.objects.select_related('curso', 'curso__grado'),
        pk=curso_docente_id,
        activo=True,
        **_attendance_filter_for_user(request.user),
    )
    alumno = get_object_or_404(Empleado, pk=alumno_id)
    detalles = list(
        AsistenciaDetalle.objects.select_related('asistencia', 'asistencia__periodo')
        .filter(asistencia__curso_docente=curso_docente, alumno=alumno)
        .order_by('-asistencia__fecha')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Historial alumno'

    style_title(ws, 1, 'Historial de asistencia por alumno')
    label_font = Font(bold=True)
    ws.cell(row=3, column=1, value='Curso:').font = label_font
    ws.cell(row=3, column=2, value=curso_docente.curso.nombre)
    ws.cell(row=4, column=1, value='Alumno:').font = label_font
    ws.cell(row=4, column=2, value=f'{alumno.apellidos}, {alumno.nombres}')

    header_row = 6
    style_table_header(ws, header_row, ['No.', 'Fecha', 'Periodo', 'Estado'])
    ws.freeze_panes = f'A{header_row + 1}'

    presentes = 0
    ausentes = 0
    for idx, detalle in enumerate(detalles, start=1):
        estado = 'Presente' if detalle.presente else 'Ausente'
        if detalle.presente:
            presentes += 1
        else:
            ausentes += 1
        style_table_row(
            ws,
            header_row + idx,
            [
                idx,
                detalle.asistencia.fecha.strftime('%d/%m/%Y'),
                detalle.asistencia.periodo.nombre if detalle.asistencia.periodo else 'Sin periodo',
                estado,
            ],
        )

    total_row = header_row + len(detalles) + 2
    ws.cell(row=total_row, column=1, value='Total presentes').font = label_font
    ws.cell(row=total_row, column=2, value=presentes)
    ws.cell(row=total_row + 1, column=1, value='Total ausentes').font = label_font
    ws.cell(row=total_row + 1, column=2, value=ausentes)

    autosize_columns(ws)
    return workbook_to_response(wb, f'historial_alumno_{alumno.id}')




@login_required
@user_passes_test(_can_view_attendance)
@require_POST
def docente_periodo_delete(request, periodo_id):
    periodo = get_object_or_404(
        PeriodoAcademico.objects.select_related('curso_docente', 'curso_docente__curso'),
        pk=periodo_id,
        **_attendance_filter_for_user(request.user, 'curso_docente__'),
    )
    curso_docente_id = periodo.curso_docente_id
    total_asistencias = periodo.asistencias.count()
    confirmed = (request.POST.get('confirm_delete') == '1')

    if total_asistencias and not confirmed:
        messages.warning(request, 'Este periodo tiene asistencias asociadas. Confirme la eliminación para continuar.')
        return redirect('empleados:docente_asistencia_home', curso_docente_id=curso_docente_id)

    if total_asistencias:
        messages.warning(request, f'Se eliminarán {total_asistencias} asistencias asociadas al periodo.')

    periodo.delete()
    messages.success(request, 'Periodo eliminado correctamente.')
    return redirect('empleados:docente_asistencia_home', curso_docente_id=curso_docente_id)


@login_required
@user_passes_test(_can_manage)
@require_GET
def buscar_alumno(request, est_id, ciclo_id, car_id, grado_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    _get_grado(est_id, ciclo_id, car_id, grado_id)

    codigo = (request.GET.get('codigo') or '').strip()
    if not codigo:
        return JsonResponse({'found': False, 'error': 'Ingrese un código personal.'}, status=400)

    qs = Empleado.objects.filter(codigo_personal__iexact=codigo)
    if not qs.exists():
        qs = Empleado.objects.filter(codigo_personal__icontains=codigo)

    alumno = qs.order_by('apellidos', 'nombres').first()
    if not alumno:
        return JsonResponse({'found': False, 'error': 'Alumno no encontrado.'}, status=404)

    return JsonResponse({
        'found': True,
        'alumno': {
            'id': alumno.id,
            'codigo': alumno.codigo_personal,
            'nombres': alumno.nombres,
            'apellidos': alumno.apellidos,
            'cui': alumno.cui,
        },
    })


@login_required
@user_passes_test(_can_manage)
@require_POST
def matricular_alumno(request, est_id, ciclo_id, car_id, grado_id):
    denied = _ensure_establecimiento_access(request, est_id)
    if denied:
        return denied
    grado = _get_grado(est_id, ciclo_id, car_id, grado_id)
    if not grado.carrera or not grado.carrera.ciclo_escolar_id:
        return JsonResponse({'ok': False, 'error': 'El grado no tiene establecimiento asociado.'}, status=400)

    establecimiento = grado.carrera.ciclo_escolar.establecimiento
    ciclo_activo = establecimiento.get_ciclo_activo()
    if not ciclo_activo:
        return JsonResponse({'ok': False, 'error': 'No hay ciclo escolar activo en este establecimiento. Activa uno para matricular.'}, status=409)

    alumno_id = (request.POST.get('alumno_id') or '').strip()
    if not alumno_id:
        return JsonResponse({'ok': False, 'error': 'Debe seleccionar un alumno.'}, status=400)

    alumno = Empleado.objects.filter(pk=alumno_id).first()
    if not alumno:
        return JsonResponse({'ok': False, 'error': 'Alumno no encontrado.'}, status=404)

    if not ALLOW_MULTI_GRADE_PER_CYCLE:
        other = Matricula.objects.filter(
            alumno=alumno,
            ciclo_escolar=ciclo_activo,
            grado__carrera__ciclo_escolar__establecimiento=establecimiento,
        ).exclude(grado=grado).exists()
        if other:
            return JsonResponse({'ok': False, 'error': 'El alumno ya está matriculado en otro grado de este establecimiento para el ciclo activo.'}, status=409)

    try:
        matricula, created = Matricula.objects.get_or_create(
            alumno=alumno,
            grado=grado,
            ciclo_escolar=ciclo_activo,
            defaults={
                'estado': 'activo',
                'ciclo': ciclo_activo.anio,
            },
        )
    except IntegrityError:
        return JsonResponse({'ok': False, 'error': 'El alumno ya está matriculado en este grado y ciclo activo.'}, status=409)

    if not created:
        if matricula.estado != 'activo':
            matricula.estado = 'activo'
            matricula.save(update_fields=['estado'])
        return JsonResponse({'ok': False, 'error': 'El alumno ya está matriculado en este grado y ciclo activo.'}, status=409)

    return JsonResponse({'ok': True, 'message': 'Alumno matriculado correctamente.'})


@login_required
@user_passes_test(_can_manage)
@require_POST
def desmatricular_alumno(request, matricula_id):
    matricula = get_object_or_404(Matricula, pk=matricula_id)
    matricula.estado = 'inactivo'
    matricula.save(update_fields=['estado'])
    messages.warning(request, 'Matrícula inactivada.')
    return redirect(request.POST.get('next') or 'empleados:establecimientos_list')
